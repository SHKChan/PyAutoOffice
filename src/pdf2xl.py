import re
import threading

import openpyxl
import pdfplumber

from MyLogger import LOGGER

class ExtractTextError(Exception):
    pass
class ExtractTextError(Exception):
    pass

class Pdf2Xl(threading.Thread):
    """_summary_

    Args:
        pdf_files(): all input pdf files path
        xl_file(): excel file path for output data
        datas(): output datas for excel
        exit_code(): exit status for convertion, initial with 'None', 0 for success, other value for fail
        progress() :convertion progress with maximum value=100

    Raises:
        self.exit_code: _description_
        Exception: _description_

    Returns:
        _type_: _description_
    """
    __slots__ = ('pdf_files', 'xl_file', 'datas',
                 'exit_code', 'progress', 'format', 'table_width')

    def __init__(self, pdf_files: str, xl_file: str, format: int) -> None:
        self.datas = list()
        self.exit_code = None
        self.format = format
        self.progress = 0   # Max value: 100%
        self.table_width = 10
        self.pdf_files = pdf_files
        self.xl_file = xl_file
        # 创建多线程 设置以保护模式启动，即主线程运行结束，子线程也停止运行
        super().__init__()
        self.setDaemon(True)

    def run(self):
        # Variable that stores the exception, if raised by someFunction
        self.exit_code = None
        self.rd_from_pdf()  # account for 90% in progress
        self.wt_to_xl()  # account for 10% in progress

    def join(self):
        threading.Thread.join(self)
        # Since join() returns in caller thread
        # we re-raise the caught exception
        # if any was caught
        if self.exit_code:
            raise self.exit_code

    def rd_from_pdf(self, mode='') -> None:
        total_last = 0
        len_pdfs = len(self.pdf_files)

        for pdf in self.pdf_files:
            with pdfplumber.open(pdf) as f:
                        
                try:
                # 提取PDF中的一页
                    for page in f.pages:

                        if 0 == self.format:
                            # 检查本页是否含有表单信息
                            # 获取Date,Project Number和Purchase/PENDING Number
                            texts = page.extract_text()

                            Date = ''
                            pat = re.compile('\d+\/\d+\/\d{4}')
                            ret = re.findall(pat, texts)
                            if(ret):
                                Date = ret[0]

                            PONO1 = ''
                            pat = re.compile('\d+\/\d+\/\d{4}\s.+\n')
                            ret = re.findall(pat, texts)
                            if(ret):
                                PONO1 = ret[0].replace(Date, '').strip()

                            PONO2 = PONO1

                            if(Date and PONO1 and PONO2):
                                ret = re.search('Item.+Amount\n', texts)
                                newBegin = ret.span()[1]
                                texts = texts[newBegin:]
                                Item = ''
                                while( not re.findall('^Total', texts)):
                                    curLine = re.findall('^.+\n', texts)[0]
                                    texts = re.sub('^.+\n', '', texts)
                                    # 检查是否为新一行
                                    ret = re.findall('^Parts\s', curLine)
                                    # 进入上一行信息录入,再进行本行行信息搜索
                                    if(ret):
                                        # 录入上一行信息
                                        if('Part' == Item):
                                            self.datas.append([])
                                            for n in range(self.table_width):
                                                self.datas[-1].append('')
                                            self.datas[-1][0] = Date
                                            self.datas[-1][1] = ''
                                            self.datas[-1][2] = PONO1
                                            self.datas[-1][3] = Item
                                            self.datas[-1][4] = PONO2
                                            self.datas[-1][5] = Description
                                            self.datas[-1][6] = Qty
                                            self.datas[-1][7] = UnitP
                                        
                                        Item = ret[0][0:-2]
                                        curLine = re.sub(ret[0], '', curLine)
                                        #逆序搜索各种信息
                                        tempLine = curLine[::-1][1:-1]
                                        tempLine = re.sub('^\d{2}\.\d+\s', '', tempLine) # 删去 Price

                                        UnitP = re.findall('^\d+\.\d+\s', tempLine)[0][::-1].strip()#获得Unit Price
                                        tempLine = re.sub('^\d+\.\d+\s', '', tempLine)

                                        Qty = re.findall('^\d+', tempLine)[0][::-1]#获得Qty
                                        tempLine = re.sub('^\d+,*\d*\s', '', tempLine)

                                        # 获取'/'前面的Description
                                        # 需要考虑'/'在下一行的情况
                                        Description = tempLine[::-1]
                                    # 否则依然位上一行信息
                                    # 若本行含有'/',则需要补充至上一行Description
                                    else:
                                        Description += curLine.strip()

                                # 录入最后一行信息
                                if('Part' == Item):
                                    self.datas.append([])
                                    for n in range(self.table_width):
                                        self.datas[-1].append('')
                                    self.datas[-1][0] = Date
                                    self.datas[-1][1] = ''
                                    self.datas[-1][2] = PONO1
                                    self.datas[-1][3] = Item
                                    self.datas[-1][4] = PONO2
                                    self.datas[-1][5] = Description
                                    self.datas[-1][6] = Qty
                                    self.datas[-1][7] = UnitP 
                    
                        elif 1 == self.format:
                            # 检查本页是否含有表单信息
                            # 获取Date,Project Number和Purchase/PENDING Number
                            texts = page.extract_text()

                            Date = ''
                            pat = re.compile('Order Date: \d+/\d+/\d{4}')
                            ret = re.findall(pat, texts)
                            if(ret):
                                Date = ret[0].replace('Order Date: ', '')

                            PONO2 = ''
                            pat = re.compile('VENDOR Vendor Quote #: .+ ')
                            ret = re.findall(pat, texts)
                            if(ret):
                                PONO2 = ret[0].replace('VENDOR Vendor Quote #: ', '')

                            PONO1 = ''
                            pat1 = re.compile('Purchase Order No.: .+')
                            ret1 = re.findall(pat1, texts)
                            pat2 = re.compile('PENDING PO No.: .+')
                            ret2 = re.findall(pat2, texts)
                            if(ret1):
                                PONO1 = ret1[0].replace('Purchase Order No.: ', '')
                            elif(ret2):
                                PONO1 = ret2[0].replace('PENDING PO No.: ', '')

                            if(Date and PONO1 and PONO2):
                                ret = re.search('Line # .+ Price', texts)
                                newBegin = ret.span()[1] + 1
                                texts = texts[newBegin:]
                                lineSharp = 1
                                while( not re.findall('^Midwest Composite', texts)):
                                    curLine = re.findall('^.+\n', texts)[0]
                                    texts = re.sub('^.+\n', '', texts)
                                    # 检查是否为新一行
                                    ret = re.findall('\d+.+\s\$\d+.+\s\$\d+.+\s\$\d+.+', curLine)
                                    # 进入上一行信息录入,再进行本行行信息搜索
                                    if(ret):
                                        lineSharp = int(re.findall('^\d+\s', curLine)[0])
                                        curLine = re.sub('^\d+\s', '', curLine)
                                        # 录入上一行信息
                                        if(lineSharp > 1 and lineSharp == len(self.datas) + 2):
                                            self.datas.append([])
                                            for n in range(self.table_width):
                                                self.datas[-1].append('')
                                            self.datas[-1][0] = Date
                                            self.datas[-1][1] = 'MTC'
                                            self.datas[-1][2] = PONO1
                                            self.datas[-1][3] = ''
                                            self.datas[-1][4] = PONO2
                                            self.datas[-1][5] = Description
                                            self.datas[-1][6] = Qty
                                            self.datas[-1][7] = UnitP
                                        #逆序搜索各种信息
                                        tempLine = curLine[::-1][1:-1]
                                        tempLine = re.sub('^\d{2}.\d+,*\d*\$\s', '', tempLine) # 删去 Price
                                        tempLine = re.sub('^\d{2}.\d+,*\d*\$\s', '', tempLine) # 删去 Discount

                                        UnitP = re.findall('^\d+.\d+,*\d*\$\s', tempLine)[0][::-1]#获得Unit Price
                                        tempLine = re.sub('^\d+.\d+,*\d*\$\s', '', tempLine)

                                        Qty = re.findall('^\d+,*\d*\s', tempLine)[0][::-1]#获得Qty
                                        tempLine = re.sub('^\d+,*\d*\s', '', tempLine)

                                        # 获取'/'前面的Description
                                        # 需要考虑'/'在下一行的情况
                                        retList = re.findall('^.+\/', curLine)
                                        if(retList):
                                            Description = retList[0][:-2]
                                        else:
                                            Description = tempLine[::-1]
                                    # 否则依然位上一行信息
                                    # 若本行含有'/',则需要补充至上一行Description
                                    else:
                                        if(re.findall('.*\s\/', curLine)):
                                            Description += re.findall('.*\s\/', curLine)[0][:-2]

                                # 录入最后一行信息
                                if(lineSharp > 0):
                                    self.datas.append([])
                                    for n in range(self.table_width):
                                        self.datas[-1].append('')
                                    self.datas[-1][0] = Date
                                    self.datas[-1][1] = 'MTC'
                                    self.datas[-1][2] = PONO1
                                    self.datas[-1][3] = ''
                                    self.datas[-1][4] = PONO2
                                    self.datas[-1][5] = Description
                                    self.datas[-1][6] = Qty
                                    self.datas[-1][7] = UnitP   
                    
                    self.progress += 1/len_pdfs*50

                except Exception as e:
                    self.exit_code = e
                    LOGGER.wt()
                    raise ExtractTextError('Error extracting text from page %d: %s' % (page.page_number, str(e)))

    def wt_to_xl(self) -> None:
        # 打开指定路径的Excel文件
        wb = openpyxl.load_workbook(self.xl_file)
        ws_names = wb.get_sheet_names()
        ws = wb.get_sheet_by_name(ws_names[0])

        # 找第一个空行
        for row in range(1, ws.max_row+2):
            for col in range(1, self.table_width+1):
                if ws.cell(row, col).value == None:
                    first_row2wt = row
                    break

        # 将数据写进sheet表单中的第一个空行
        len_datas = len(self.datas)
        for row in range(len_datas):
            for col in range(self.table_width):
                ws.cell(row+first_row2wt, col+1).value = self.datas[row][col]
            self.progress += 1/len_datas*50

        # Save the Excel workbook, retrying if a PermissionError occurs
        try:
            wb.save(self.xl_file)
            self.exit_code = 0
        except PermissionError as e:
            self.exit_code = e
            LOGGER.wt()
            raise PermissionError('File is currently open, retrying...')
        
        wb.close()
